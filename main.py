from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import time
from datetime import datetime, timedelta
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import TimeoutException, StaleElementReferenceException
from selenium.webdriver.common.action_chains import ActionChains
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv
import os

load_dotenv()
USUARIO = os.getenv("USUARIO_MASTER")
SENHA = os.getenv("SENHA_MASTER")
print(f"USUARIO: {USUARIO}")
print(f"SENHA: {SENHA}")
navegador = webdriver.Chrome()
navegador.get("https://autenticacao.bancomaster.com.br/login")
navegador.maximize_window()

# Preencher login
print("Preenchendo usuario...")
navegador.find_element(
    By.XPATH, '//*[@id="mat-input-0"]').send_keys(USUARIO)
print("Preenchendo senha...")
navegador.find_element(By.XPATH, '//*[@id="mat-input-1"]').send_keys(SENHA)

# Clicar no botão "Entrar"
print("Entrando...")
botao_entrar = navegador.find_element(
    By.XPATH, '/html/body/app-root/app-login/div/div[2]/mat-card/mat-card-content/form/div[3]/button[2]')
botao_entrar.click()

# Tratamento para a tela de "já conectado"
try:
    botao_sim = WebDriverWait(navegador, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, '//*[@id="mat-dialog-0"]/app-confirmacao-dialog/div/div[3]/div/app-botao-icon-v2[2]/button'))
    )
    botao_sim.click()
    print("Conexão anterior confirmada com 'Sim'.")
except:
    print("Tela de conexão anterior não apareceu.")


# Clicar no item menu_tetris
menu_tetris = WebDriverWait(navegador, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH, "//i[contains(@class, 'icon-element-3')]"))
)
print("clicando no tetris...")
menu_tetris.click()

# Esperar que o modal ou o conteúdo relacionado tenha sido carregado completamente
# Isso pode incluir esperar até que o modal ou um elemento específico dentro dele esteja visível
time.sleep(2)
modal_esperado = WebDriverWait(navegador, 10).until(
    # Verifica se o elemento do modal está visível
    EC.visibility_of_element_located(
        (By.XPATH, '//*[@id="dorp_list_icons_0"]'))
)
print("Esperando modal...")
# Agora, clicar no elemento dentro do modal
vendas_consig = WebDriverWait(navegador, 10).until(
    EC.element_to_be_clickable((By.XPATH, '//*[@id="dorp_list_icons_0"]'))
)
print("clicando em vendas consignadas...")
vendas_consig.click()

time.sleep(5)
menu_hamburguer = WebDriverWait(navegador, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH, '/html/body/app-root/app-home-v2/app-master-header-layout/mat-toolbar/button/span[1]/mat-icon'))
)
print("Apertando menu hamburguer...")
menu_hamburguer.click()

time.sleep(2)

cadastro_saque = WebDriverWait(navegador, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="mat-expansion-panel-header-0"]/span/mat-panel-title'))
)
print("clicando cadastro saque...")
cadastro_saque.click()

botao_consul_esteira = WebDriverWait(navegador, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="cdk-accordion-child-0"]/div/mat-list-item[2]/span/a/span[1]'))
)
print("clicando botao consultar esteira...")
botao_consul_esteira.click()


# Obtém a data de hoje (data final)
data_final = datetime.now()
data_inicial = data_final - timedelta(days=30)
data_inicial_formatada = data_inicial.strftime('%d/%m/%Y')
data_final_formatada = data_final.strftime('%d/%m/%Y')
campo_input_esquerda = WebDriverWait(navegador, 20).until(
    EC.visibility_of_element_located(
        (By.XPATH, "//input[contains(@class, 'mat-start-date')]"))
)

campo_input_esquerda.clear()
campo_input_esquerda.send_keys(data_inicial_formatada)


campo_input_direita = WebDriverWait(navegador, 20).until(
    EC.element_to_be_clickable(
        (By.XPATH, "//input[contains(@class, 'mat-end-date')]"))
)

campo_input_direita.clear()
campo_input_direita.send_keys(data_final_formatada)
print("Data selecionada...")


botao_status_forma = WebDriverWait(navegador, 5).until(
    EC.element_to_be_clickable(
        (By.XPATH, '//*[@id="mat-select-value-7"]/span'))
)
print("selecionando status forma...")
botao_status_forma.click()

status_forma = WebDriverWait(navegador, 20).until(
    EC.visibility_of_element_located(
        (By.XPATH, "//span[text()=' Aguardando Aprovação Corban ']")
    )
)
print("clicando status...")
status_forma.click()

# Localiza o botão "Buscar" e clica nele
botao_buscar = WebDriverWait(navegador, 20).until(
    EC.element_to_be_clickable(
        (By.XPATH, '/html/body/app-root/app-home-v2/app-side-bar/mat-sidenav-container/mat-sidenav-content/mat-card/app-saque-esteira/mat-card-content/app-form-saque-consulta-esteira/div/form/app-botao-icon-v2/button')
    )
)
print("apertando botao buscar...")
botao_buscar.click()

# Aguarda por um elemento que indique ausência de resultados
try:
    mensagem_sem_resultado = WebDriverWait(navegador, 40).until(
        EC.presence_of_element_located(
            (By.XPATH, "//span[contains(text(), 'Nenhum arquivo encontrado')]")
        )
    )
    print("Nenhum arquivo encontrado. Encerrando a aplicação.")
    navegador.quit()
except TimeoutException:
    print("Arquivos encontrados ou a mensagem de erro não apareceu. Continuando o script.")


tentativas = 3

for tentativa in range(tentativas):
    try:
        botao_exportar = WebDriverWait(navegador, 60).until(
            EC.element_to_be_clickable(
                (By.XPATH, '/html/body/app-root/app-home-v2/app-side-bar/mat-sidenav-container/mat-sidenav-content/mat-card/app-saque-esteira/mat-card-content/app-form-saque-consulta-esteira/div/app-botao-icon-v2/button')
            )
        )
        print("Botão Exportar clicado!")
        botao_exportar.click()
        break
    except TimeoutException:
        print(f"Tentativa {tentativa +
              1} para clicar no botão Exportar... continuando.")
        if tentativa == tentativas - 1:
            print(
                "Botão Exportar não encontrado após 3 tentativas. Encerrando o navegador.")
            navegador.quit()


##########################################
# Clicar no item menu_tetris
menu_tetris = WebDriverWait(navegador, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH, "//i[contains(@class, 'icon-element-3')]"))
)
print("clicando no tetris...")
menu_tetris.click()


time.sleep(2)
modal_esperado = WebDriverWait(navegador, 10).until(
    # Verifica se o elemento do modal está visível
    EC.visibility_of_element_located(
        (By.XPATH, '//*[@id="dorp_list_icons_0"]'))
)
print("Esperando modal...")
# Agora, clicar no elemento dentro do modal
relatorio_menu = WebDriverWait(navegador, 10).until(
    EC.presence_of_element_located(
        (By.XPATH,
         "//*[@id='dorp_list_icons_1']//mat-icon[@data-mat-icon-name='task-square-bm']"))
)

print("clicando em vendas consignadas...")
relatorio_menu.click()

time.sleep(2)
menu_hamburguer = WebDriverWait(navegador, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH, '/html/body/app-root/app-home-v2/app-master-header-layout/mat-toolbar/button/span[1]/mat-icon'))
)
print("Apertando menu hamburguer...")
menu_hamburguer.click()

relatorio_menu_hamburguer = WebDriverWait(navegador, 10).until(
    EC.presence_of_element_located(
        (By.XPATH, "//mat-panel-title[contains(text(), 'Relatórios')]"))
)
print("Clicando relatorio dentro do menu...")
relatorio_menu_hamburguer.click()


botao_relatorio_esteira = WebDriverWait(navegador, 10).until(
    EC.element_to_be_clickable(
        (By.XPATH,
         "//mat-list-item//a[contains(@href, '/relatorio/esteira-saque')]"))
)
print("Apertando botão 'Relatório Esteira Saque'...")
botao_relatorio_esteira.click()

##############################

try:
    # Espera pelo botão de download e clica
    botao_relatorio_download = WebDriverWait(navegador, 10).until(
        EC.element_to_be_clickable(
            (By.XPATH, "//button[.//mat-icon[text()='download']]")
        )
    )
    print("Apertando botão 'download'...")
    botao_relatorio_download.click()

    # Aguarda um tempo para o download ser concluído
    print("Aguardando o download ser concluído...")
    # Tempo para garantir que o download ocorra (ajuste conforme necessário)
    time.sleep(10)

except TimeoutException:
    print("Botão 'download' não encontrado. Fechando o navegador.")
    navegador.quit()  # Fecha o navegador se o botão não for encontrado
    exit()  # Encerra o script
print("Download (ou tentativa de download) concluído. Fechando a automação.")

navegador.quit()  # Fecha o navegador ao final

# Caminho da pasta de downloads
pasta_downloads = os.path.expanduser("~/Downloads")

# Função para encontrar o arquivo mais recente
def encontrar_arquivo_mais_recente(pasta, extensao=".xlsx"):
    try:
        arquivos = [
            os.path.join(pasta, f)
            for f in os.listdir(pasta)
            if os.path.isfile(os.path.join(pasta, f)) and f.endswith(extensao) and not f.startswith('~$')
        ]
        if not arquivos:
            print("Nenhum arquivo Excel encontrado na pasta de downloads.")
            return None

        arquivo_mais_recente = max(arquivos, key=os.path.getctime)
        print(f"Arquivo mais recente encontrado: {arquivo_mais_recente}")
        return arquivo_mais_recente
    except Exception as e:
        print(f"Erro ao localizar o arquivo mais recente: {e}")
        return None

# Função para verificar e ajustar o arquivo e colunas
def verificar_e_ajustar_headers(arquivo, headers_esperados):
    try:
        tabela = pd.read_excel(arquivo)
        headers_do_arquivo = list(tabela.columns)

        # Verifica colunas faltantes
        colunas_faltantes = [h for h in headers_esperados if h not in headers_do_arquivo]
        if colunas_faltantes:
            print("As seguintes colunas estão faltando no arquivo:")
            for coluna in colunas_faltantes:
                print(f"- {coluna}")

        # Remove colunas que não estão na lista esperada
        tabela_ajustada = tabela[[col for col in headers_do_arquivo if col in headers_esperados]]

        # Salva arquivo corrigido
        arquivo_corrigido = arquivo.replace(".xlsx", "_corrigido.xlsx")
        tabela_ajustada.to_excel(arquivo_corrigido, index=False)
        print(f"Arquivo corrigido salvo como: {arquivo_corrigido}")
        return arquivo_corrigido

    except Exception as e:
        print(f"Erro ao carregar ou verificar o arquivo: {e}")
        return None

# Função para enviar os dados ao Google Sheets
def atualizar_google_sheets(arquivo, credenciais, url_planilha, aba):
    try:
        print(f"Iniciando atualização do Google Sheets com o arquivo: {arquivo}")

        # Configurar as credenciais e acessar o Google Sheets
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(credenciais, scope)
        client = gspread.authorize(creds)

        print("Credenciais autorizadas com sucesso.")

        # Abrir a planilha e a aba específica
        sheet = client.open_by_url(url_planilha).worksheet(aba)
        print("Planilha do Google Sheets aberta com sucesso.")

        # Ler o arquivo corrigido
        tabela = pd.read_excel(arquivo)

        # Substituir NaN por valores vazios antes de inserir os dados
        tabela.fillna('', inplace=True)

        # Inserir dados a partir da segunda linha
        new_data = tabela.values.tolist()
        sheet.insert_rows(new_data, 2)

        print("Dados enviados para o Google Sheets com sucesso.")

        # Atualizar colunas "MASTER" e "NÃO" com fórmulas
        master_index = headers_esperados.index("MASTER")  # Índice da coluna MASTER
        nao_index = headers_esperados.index("NÃO")  # Índice exato da coluna "NÃO"
        nao_pertube_index = headers_esperados.index("NÃO PERTUBE") if "NÃO PERTUBE" in headers_esperados else None
        start_row = 2
        end_row = start_row + len(new_data) - 1

        # Atualizar coluna MASTER
        master_column = get_column_letter(master_index + 1)
        master_range = f'{master_column}{start_row}:{master_column}{end_row}'
        master_values = [['MASTER']] * len(new_data)
        sheet.update(master_range, master_values)

        # Limpar coluna "NÃO PERTUBE"
        if nao_pertube_index is not None:
            nao_pertube_column = get_column_letter(nao_pertube_index + 1)
            nao_pertube_range = f'{nao_pertube_column}{start_row}:{nao_pertube_column}{end_row}'
            empty_values = [['']] * len(new_data)
            sheet.update(nao_pertube_range, empty_values)

        # Atualizar coluna "NÃO" com fórmula
        formula = 'SEERRO(SE(E(CONT.SE(GERAL!$A$2:A;A{row})>0;PROCV(A{row};GERAL!$A$2:P;16;0)="MASTER");"SIM";"NÃO");"NÃO")'
        nao_column = 'AY'  # A coluna AY corresponde à 51ª coluna

        for row in range(start_row, end_row + 1):
            cell_range = f'{nao_column}{row}'
            formula_value = formula.format(row=row)
            sheet.update_acell(cell_range, f'={formula_value}')

        print("Colunas MASTER e NÃO atualizadas com sucesso.")

        # Excluir linhas com "SIM" na coluna "NÃO"
        nao_values = sheet.col_values(nao_index + 1)
        rows_to_delete = [idx for idx, value in enumerate(nao_values[1:], start=2) if value == "SIM"]
        for row in reversed(rows_to_delete):
            sheet.delete_rows(row)

        print("Linhas com SIM na coluna NÃO foram excluídas.")

        # Excluir o arquivo após o upload
        os.remove(arquivo)
        print(f"Arquivo {arquivo} excluído com sucesso.")

    except Exception as e:
        print(f"Erro ao atualizar o Google Sheets: {e}")

# Executa a automação
print("Executando a automação...")

headers_esperados = [
    "STATUS FORMALIZAÇÃO", "CONVÊNIO", "DIA DE CORTE", "TIPO", "CPF", "NOME", "MATRÍCULA", "NSU", "PROPOSTA",
    "VALOR SAQUE", "VALOR TROCO", "VALOR PARCELA", "PRAZO", "BANCO", "AGÊNCIA", "CONTA", "DIGITADOR", "DATA SAQUE",
    "CANAL VENDA", "PENDENTE DADOS BANCÁRIOS", "CRÍTICA", "DATA DE APROVAÇÃO", "USUÁRIO APROVAÇÃO", "ID CONVENIO",
    "TIPO AUDITORIA DIGITAL", "STATUS AUDITORIA DIGITAL", "VENDA PACOTE VANTAGENS", "ACEITE PACOTE VANTAGENS",
    "VALOR PREMIO LIQUIDO", "VALOR PREMIO BRUTO", "STATUS PROPOSTA FUNCAO", "ATIVIDADE", "ORIGEM", "STATUS TED",
    "DATA EMISSÃO TED", "ÚLTIMO EVENTO TED", "MOTIVO FUNÇÃO", "TABELA FUNÇÃO", "POSSUI REPRESENTANTE LEGAL",
    "CPF REPRESENTANTE LEGAL", "NOME REPRESENTANTE LEGAL", "DIGITADOR", "COD CORRESPONDENTE", "PRODUTO",
    "TIPO_COBRANCA", "NUM PARCELAS PREMIO", "VALOR PARCELA PREMIO", "NÃO", "FORNECEDOR BIOMETRIA",
    "MASTER", "NÃO PERTUBE"  # Adicionando cabeçalhos esperados
]

arquivo_mais_recente = encontrar_arquivo_mais_recente(pasta_downloads)

if arquivo_mais_recente:
    arquivo_corrigido = verificar_e_ajustar_headers(arquivo_mais_recente, headers_esperados)
    if arquivo_corrigido:
        print("Arquivo processado com sucesso. Enviando ao Google Sheets...")

        # Configurações do Google Sheets
        credenciais_path = r'C:\Users\mateu\banco_master\credentials.json'
        planilha_url = "https://docs.google.com/spreadsheets/d/1mcucTR78B5uX3c4MPgMP65rFN3PxF3-UltNCkDoFrnw/edit?usp=sharing"
        aba = "MASTER"  # Certifique-se de que a aba correta está sendo selecionada

        atualizar_google_sheets(arquivo_corrigido, credenciais_path, planilha_url, aba)
    else:
        print("Erro ao processar o arquivo. Verifique os detalhes acima.")
else:
    print("Erro: Nenhum arquivo Excel encontrado. Encerrando a aplicação.")

print("Automação encerrada.")
