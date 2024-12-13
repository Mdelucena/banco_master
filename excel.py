import os
import pandas as pd
import gspread
from oauth2client.service_account import ServiceAccountCredentials
from openpyxl.utils import get_column_letter

# Caminho da pasta de downloads
pasta_downloads = os.path.expanduser("~/Downloads")

# Função para encontrar o arquivo mais recente
def encontrar_arquivo_mais_recente(pasta, extensao=".xlsx"):
    try:
        arquivos = [
            os.path.join(pasta, f)
            for f in os.listdir(pasta)
            if os.path.isfile(os.path.join(pasta, f)) and f.endswith(extensao) and not f.startswith('~$')
        ]
        if not arquivos:
            print("Nenhum arquivo Excel encontrado na pasta de downloads.")
            return None

        arquivo_mais_recente = max(arquivos, key=os.path.getctime)
        print(f"Arquivo mais recente encontrado: {arquivo_mais_recente}")
        return arquivo_mais_recente
    except Exception as e:
        print(f"Erro ao localizar o arquivo mais recente: {e}")
        return None

# Função para verificar e ajustar o arquivo e colunas
def verificar_e_ajustar_headers(arquivo, headers_esperados):
    try:
        tabela = pd.read_excel(arquivo)
        headers_do_arquivo = list(tabela.columns)

        # Verifica colunas faltantes
        colunas_faltantes = [h for h in headers_esperados if h not in headers_do_arquivo]
        if colunas_faltantes:
            print("As seguintes colunas estão faltando no arquivo:")
            for coluna in colunas_faltantes:
                print(f"- {coluna}")

        # Remove colunas que não estão na lista esperada
        tabela_ajustada = tabela[[col for col in headers_do_arquivo if col in headers_esperados]]

        # Salva arquivo corrigido
        arquivo_corrigido = arquivo.replace(".xlsx", "_corrigido.xlsx")
        tabela_ajustada.to_excel(arquivo_corrigido, index=False)
        print(f"Arquivo corrigido salvo como: {arquivo_corrigido}")
        return arquivo_corrigido

    except Exception as e:
        print(f"Erro ao carregar ou verificar o arquivo: {e}")
        return None

# Função para enviar os dados ao Google Sheets
def atualizar_google_sheets(arquivo, credenciais, url_planilha, aba):
    try:
        print(f"Iniciando atualização do Google Sheets com o arquivo: {arquivo}")

        # Configurar as credenciais e acessar o Google Sheets
        scope = ['https://spreadsheets.google.com/feeds', 'https://www.googleapis.com/auth/drive']
        creds = ServiceAccountCredentials.from_json_keyfile_name(credenciais, scope)
        client = gspread.authorize(creds)

        print("Credenciais autorizadas com sucesso.")

        # Abrir a planilha e a aba específica
        sheet = client.open_by_url(url_planilha).worksheet(aba)
        print("Planilha do Google Sheets aberta com sucesso.")

        # Ler o arquivo corrigido
        tabela = pd.read_excel(arquivo)

        # Substituir NaN por valores vazios antes de inserir os dados
        tabela.fillna('', inplace=True)

        # Inserir dados a partir da segunda linha
        new_data = tabela.values.tolist()
        sheet.insert_rows(new_data, 2)

        print("Dados enviados para o Google Sheets com sucesso.")

        # Atualizar colunas "MASTER" e "NÃO" com fórmulas
        master_index = headers_esperados.index("MASTER")  # Índice da coluna MASTER
        nao_index = headers_esperados.index("NÃO")  # Índice exato da coluna "NÃO"
        nao_pertube_index = headers_esperados.index("NÃO PERTUBE") if "NÃO PERTUBE" in headers_esperados else None
        start_row = 2
        end_row = start_row + len(new_data) - 1

        # Atualizar coluna MASTER
        master_column = get_column_letter(master_index + 1)
        master_range = f'{master_column}{start_row}:{master_column}{end_row}'
        master_values = [['MASTER']] * len(new_data)
        sheet.update(master_range, master_values)

        # Limpar coluna "NÃO PERTUBE"
        if nao_pertube_index is not None:
            nao_pertube_column = get_column_letter(nao_pertube_index + 1)
            nao_pertube_range = f'{nao_pertube_column}{start_row}:{nao_pertube_column}{end_row}'
            empty_values = [['']] * len(new_data)
            sheet.update(nao_pertube_range, empty_values)

        # Atualizar coluna "NÃO" com fórmula
        formula = 'SEERRO(SE(E(CONT.SE(GERAL!$A$2:A;A{row})>0;PROCV(A{row};GERAL!$A$2:P;16;0)="MASTER");"SIM";"NÃO");"NÃO")'
        nao_column = 'AY'  # A coluna AY corresponde à 51ª coluna

        for row in range(start_row, end_row + 1):
            cell_range = f'{nao_column}{row}'
            formula_value = formula.format(row=row)
            sheet.update_acell(cell_range, f'={formula_value}')

        print("Colunas MASTER e NÃO atualizadas com sucesso.")

        # Excluir linhas com "SIM" na coluna "NÃO"
        nao_values = sheet.col_values(nao_index + 1)
        rows_to_delete = [idx for idx, value in enumerate(nao_values[1:], start=2) if value == "SIM"]
        for row in reversed(rows_to_delete):
            sheet.delete_rows(row)

        print("Linhas com SIM na coluna NÃO foram excluídas.")

        # Excluir o arquivo após o upload
        os.remove(arquivo)
        print(f"Arquivo {arquivo} excluído com sucesso.")

    except Exception as e:
        print(f"Erro ao atualizar o Google Sheets: {e}")

# Executa a automação
print("Executando a automação...")

headers_esperados = [
    "STATUS FORMALIZAÇÃO", "CONVÊNIO", "DIA DE CORTE", "TIPO", "CPF", "NOME", "MATRÍCULA", "NSU", "PROPOSTA",
    "VALOR SAQUE", "VALOR TROCO", "VALOR PARCELA", "PRAZO", "BANCO", "AGÊNCIA", "CONTA", "DIGITADOR", "DATA SAQUE",
    "CANAL VENDA", "PENDENTE DADOS BANCÁRIOS", "CRÍTICA", "DATA DE APROVAÇÃO", "USUÁRIO APROVAÇÃO", "ID CONVENIO",
    "TIPO AUDITORIA DIGITAL", "STATUS AUDITORIA DIGITAL", "VENDA PACOTE VANTAGENS", "ACEITE PACOTE VANTAGENS",
    "VALOR PREMIO LIQUIDO", "VALOR PREMIO BRUTO", "STATUS PROPOSTA FUNCAO", "ATIVIDADE", "ORIGEM", "STATUS TED",
    "DATA EMISSÃO TED", "ÚLTIMO EVENTO TED", "MOTIVO FUNÇÃO", "TABELA FUNÇÃO", "POSSUI REPRESENTANTE LEGAL",
    "CPF REPRESENTANTE LEGAL", "NOME REPRESENTANTE LEGAL", "DIGITADOR", "COD CORRESPONDENTE", "PRODUTO",
    "TIPO_COBRANCA", "NUM PARCELAS PREMIO", "VALOR PARCELA PREMIO", "NÃO", "FORNECEDOR BIOMETRIA",
    "MASTER", "NÃO PERTUBE"  # Adicionando cabeçalhos esperados
]

arquivo_mais_recente = encontrar_arquivo_mais_recente(pasta_downloads)

if arquivo_mais_recente:
    arquivo_corrigido = verificar_e_ajustar_headers(arquivo_mais_recente, headers_esperados)
    if arquivo_corrigido:
        print("Arquivo processado com sucesso. Enviando ao Google Sheets...")

        # Configurações do Google Sheets
        credenciais_path = r'C:\Users\mateu\banco_master\credentials.json'
        planilha_url = "https://docs.google.com/spreadsheets/d/1mcucTR78B5uX3c4MPgMP65rFN3PxF3-UltNCkDoFrnw/edit?usp=sharing"
        aba = "MASTER"  # Certifique-se de que a aba correta está sendo selecionada

        atualizar_google_sheets(arquivo_corrigido, credenciais_path, planilha_url, aba)
    else:
        print("Erro ao processar o arquivo. Verifique os detalhes acima.")
else:
    print("Erro: Nenhum arquivo Excel encontrado. Encerrando a aplicação.")

print("Automação encerrada.")